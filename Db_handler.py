import sqlite3
from datetime import datetime
import openpyxl
import os

DB_NAME = "RECD.db"

def init_db():
    """Initializes the SQLite database and tables."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS machines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                channel_id TEXT NOT NULL,
                api_key TEXT NOT NULL
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS machine_controllers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                machine_id INTEGER,
                controller TEXT,
                controller_no TEXT,
                mfg_date TEXT,
                inst_date TEXT,
                customer_name TEXT,
                FOREIGN KEY(machine_id) REFERENCES machines(id) ON DELETE CASCADE
            )
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS fault_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                machine_name TEXT NOT NULL,
                vm_field TEXT NOT NULL,
                hex_value TEXT NOT NULL
            )
        """)

        # ✅ Admin table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        """)
        
        conn.commit()
    print("[INIT] Database initialized with required tables.")

def has_any_admin():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        # ✅ Safe check for table existence
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='admins'")
        if not cursor.fetchone():
            return False
        cursor.execute("SELECT 1 FROM admins LIMIT 1")
        return cursor.fetchone() is not None

def create_admin(username, password):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO admins (username, password) VALUES (?, ?)", (username, password))
        conn.commit()

def get_admin(username, password):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM admins WHERE username=? AND password=?", (username, password))
        return cursor.fetchone() is not None

def reset_admin(username, new_password):
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE admins SET password=? WHERE username=?", (new_password, username))
        if cursor.rowcount == 0:
            return False
        conn.commit()
        return True

def list_admins():
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT username FROM admins")
        return [row[0] for row in cursor.fetchall()]

def add_machine(name, channel_id, api_key, controller_data):
    """Add a new machine and its controller data."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM machines WHERE name = ?", (name,))
        if cursor.fetchone():
            raise ValueError("Machine already stored!")

        cursor.execute("""
            INSERT INTO machines (name, channel_id, api_key)
            VALUES (?, ?, ?)
        """, (name, channel_id, api_key))
        machine_id = cursor.lastrowid

        for entry in controller_data:
            cursor.execute("""
                INSERT INTO machine_controllers (
                    machine_id, controller, controller_no, mfg_date, inst_date, customer_name
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                machine_id,
                entry.get("controller"),
                entry.get("controller_no"),
                entry.get("mfg_date", "N/A"),
                entry.get("inst_date", "N/A"),
                entry.get("customer_name", "N/A")
            ))
        conn.commit()

def update_machine_full(old_name, new_name, old_channel, new_channel, old_key, new_key, controller_updates):
    """Update machine and controller details."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE machines
            SET name = ?, channel_id = ?, api_key = ?
            WHERE name = ? AND channel_id = ? AND api_key = ?
        """, (new_name, new_channel, new_key, old_name, old_channel, old_key))

        if cursor.rowcount == 0:
            return False

        cursor.execute("SELECT id FROM machines WHERE name = ?", (new_name,))
        row = cursor.fetchone()
        if not row:
            return False
        machine_id = row[0]

        for item in controller_updates:
            ctrl = item["controller"]
            old_no = item["old_no"]
            new_no = item["new_no"]
            cust_name = item.get("customer_name", "N/A")

            if old_no and new_no:
                cursor.execute("""
                    UPDATE machine_controllers
                    SET controller_no = ?, customer_name = ?
                    WHERE machine_id = ? AND controller = ? AND controller_no = ?
                """, (new_no, cust_name, machine_id, ctrl, old_no))

        conn.commit()
        return True

def delete_machine(name):
    """Delete machine and related controllers."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM machines WHERE name = ?", (name,))
        row = cursor.fetchone()
        if not row:
            return
        machine_id = row[0]

        cursor.execute("DELETE FROM machine_controllers WHERE machine_id = ?", (machine_id,))
        cursor.execute("DELETE FROM machines WHERE id = ?", (machine_id,))
        conn.commit()

def insert_fault_log(machine_name, vm_field, hex_value):
    """Insert a fault log entry."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO fault_logs (timestamp, machine_name, vm_field, hex_value)
            VALUES (?, ?, ?, ?)
        """, (datetime.now().isoformat(), machine_name, vm_field, hex_value))
        conn.commit()

def load_machines():
    """Load machines and M1 controller data as summary."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, channel_id, api_key FROM machines")
        machine_rows = cursor.fetchall()

        machines = {}
        for machine_id, name, channel_id, api_key in machine_rows:
            cursor.execute("""
                SELECT controller_no, mfg_date, inst_date
                FROM machine_controllers
                WHERE machine_id = ? AND controller = 'M1'
                LIMIT 1
            """, (machine_id,))
            ctrl_data = cursor.fetchone()

            if ctrl_data:
                controller_no, mfg_date, inst_date = ctrl_data
            else:
                controller_no, mfg_date, inst_date = "N/A", "N/A", "N/A"

            machines[name] = {
                "channel_id": channel_id,
                "api_key": api_key,
                "field": 1,
                "controller_no": controller_no,
                "mfg_date": mfg_date,
                "install_date": inst_date
            }
        return machines

def insert_from_excel(file_path):
    """
    Import data from Excel where each row corresponds to one controller of one machine.
    Expected columns (case and space insensitive): 
        - "Machine Name"
        - "Channel ID"
        - "Controller Number"
        - "API Key"
        - "Fields(M1 to M8)" (the number 1 to 8 indicating controller name M1 to M8)
        - "Mfg Date"      # <-- NEW, Optional
        - "Inst Date"     # <-- NEW, Optional
        - "Customer Name" # <-- NEW, Optional
    This function is Windows encoding safe and compatible.
    """
    if not os.path.exists(file_path):
        print("Excel file not found:", file_path)
        return

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Failed to open Excel file: {e}")
        return

    sheet = wb.active

    try:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {h.lower().replace(" ", ""): i for i, h in enumerate(headers)}
    except Exception as e:
        print(f"Error reading headers: {e}")
        return

    # Prepare all required headers, match ignoring case AND spaces
    def colidx(colname):
        key = colname.lower().replace(" ", "")
        return header_map[key] if key in header_map else None

    required_columns = [
        "machine name", "channel id", "controller number", "api key", "fields(m1 to m8)"
    ]
    missing_cols = [col for col in required_columns if colidx(col) is None]
    if missing_cols:
        print(f"Missing required columns in Excel: {missing_cols}")
        print(f"Found columns: {headers}")
        return

    machine_idx   = colidx("machine name")
    channel_idx   = colidx("channel id")
    ctrl_num_idx  = colidx("controller number")
    api_idx       = colidx("api key")
    fieldno_idx   = colidx("fields(m1 to m8)")

    mfg_date_idx    = colidx("mfg date")
    inst_date_idx   = colidx("inst date")
    customer_idx    = colidx("customer name")

    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                name         = str(row[machine_idx]).strip()   if row[machine_idx]   else ""
                channel_id   = str(row[channel_idx]).strip()   if row[channel_idx]   else ""
                controller_no= str(row[ctrl_num_idx]).strip()  if row[ctrl_num_idx]  else ""
                api_key      = str(row[api_idx]).strip()       if row[api_idx]       else ""
                field_no     = str(row[fieldno_idx]).strip()   if row[fieldno_idx]   else ""

                mfg_date     = str(row[mfg_date_idx]).strip()  if (mfg_date_idx is not None and row[mfg_date_idx])   else "N/A"
                inst_date    = str(row[inst_date_idx]).strip() if (inst_date_idx is not None and row[inst_date_idx]) else "N/A"
                customer_name= str(row[customer_idx]).strip()  if (customer_idx is not None and row[customer_idx])  else "N/A"

                if not (name and channel_id and controller_no and api_key and field_no):
                    continue  # skip incomplete rows

                controller = f"M{field_no}"  # converts field_no "1" -> M1, etc.

                # Insert machine if not exists
                cursor.execute(
                    "INSERT OR IGNORE INTO machines (name, channel_id, api_key) VALUES (?, ?, ?)",
                    (name, channel_id, api_key)
                )
                cursor.execute("SELECT id FROM machines WHERE name = ?", (name,))
                machine_row = cursor.fetchone()
                if not machine_row:
                    continue
                machine_id = machine_row[0]

                # Insert or update controller info
                cursor.execute(
                    "SELECT id FROM machine_controllers WHERE machine_id=? AND controller=?",
                    (machine_id, controller)
                )
                if cursor.fetchone():
                    cursor.execute("""
                        UPDATE machine_controllers SET 
                            controller_no = ?, 
                            mfg_date = ?, 
                            inst_date = ?, 
                            customer_name = ?
                        WHERE machine_id=? AND controller=?
                        """,
                        (controller_no, mfg_date, inst_date, customer_name, machine_id, controller)
                    )
                else:
                    cursor.execute("""
                        INSERT INTO machine_controllers (
                            machine_id, controller, controller_no, mfg_date, inst_date, customer_name
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    """, (machine_id, controller, controller_no, mfg_date, inst_date, customer_name))
            except Exception as row_e:
                print(f"Failed to process row {row}: {row_e}")

        conn.commit()
    print("Excel import complete - all controllers and machines successfully imported.")
